import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import csv
import datetime
import os
import re
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill
import glob
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np


# ヘッドレスモードでブラウザを起動
options = Options()
options.add_argument('--headless')
driver = webdriver.Chrome('./driver/chromedriver', options=options)


driver.set_window_size('1200', '1000')
driver.get('https://en-gage.net/user/#/')
time.sleep(1)

area_btn = driver.find_element(By.LINK_TEXT, "選択")
area_btn.click()
time.sleep(1)

kanto_btn = driver.find_element(By.LINK_TEXT, "関東")
kanto_btn.click()
time.sleep(1)

tokyo_btn = driver.find_element(By.LINK_TEXT, "東京都")
tokyo_btn.click()
time.sleep(1)

shibuya_btn = driver.find_element(By.LINK_TEXT, "渋谷区")
shibuya_btn.click()
time.sleep(1)

option_btn = driver.find_element(By.LINK_TEXT, "職種、給与など、こだわりは？")
option_btn.click()
time.sleep(1)

checkbox_ce3 = driver.find_element(By.XPATH, "//label[@for='ce_3']")
checkbox_ce3.click()

checkbox_ce5 = driver.find_element(By.XPATH, "//label[@for='ce_5']")
checkbox_ce5.click()

checkbox_ce7 = driver.find_element(By.XPATH, "//label[@for='ce_7']")
checkbox_ce7.click()

occupation_btn = driver.find_element(By.LINK_TEXT, "選択してください")
occupation_btn.click()
time.sleep(1)

occupation_accordion = driver.find_element(By.XPATH, "//div[@class='md_accordion'][7]")
occupation_accordion.click()

checkbox_p_401000 = driver.find_element(By.XPATH, "//label[@for='p_401000']")
checkbox_p_401000.click()

checkbox_p_402000 = driver.find_element(By.XPATH, "//label[@for='p_402000']")
checkbox_p_402000.click()

# checkbox_p_409000 = driver.find_element_by_xpath("//label[@for='p_409000']")
# checkbox_p_409000.click()

choose_btn = driver.find_element(By.LINK_TEXT, "選択")
choose_btn.click()
time.sleep(1)

search_btn = driver.find_element(By.LINK_TEXT, "この条件で探す")
search_btn.click()
time.sleep(1)

# スクレイピング結果を書き出すcsvの用意
csv_date = datetime.datetime.today().strftime("%Y%m%d%H%M")
csv_file_name = 'engage' + csv_date + '.csv'

# 有効性を確認するキーワードの設定
kw01 = "リモート"
kw02 = "未経験"
kw03 = "歓迎"
kw04 = "大手"
kw05 = "休日"

# ファイルの閉じ忘れが怖いのでwithで開く
with open(csv_file_name, 'w', encoding='cp932', errors='ignore') as f:
  writer = csv.writer(f, lineterminator='\n')
  csv_header = ["検索順位", "求人タイトル", "いいね数" , "URL", kw01, kw02, kw03, kw04, kw05,]
  writer.writerow(csv_header)

  # 検索順位の定義
  item = 1

  # ループ用の変数を用意
  i = 0

  while True:
    i = i + 1
    time.sleep(1)

    # 求人のタイトルをすべて取得する
    for elem_ttl in driver.find_elements(By.XPATH, "//a[@class='headArea']/div[@class='catch']"):
      elem_a = elem_ttl.find_element(By.XPATH, "..")
      
      csvlist = []
      csvlist.append(str(item))
      csvlist.append(elem_ttl.text)

      # いいね表示の有無を判定
      if elem_a.find_elements(By.XPATH, "..//span[@class='num']"):
        
        # いいね数が表示されている場合はその数を取得
        like = elem_a.find_element(By.XPATH, "..//span[@class='num']").text
        
        # 数字以外の部分が不要なので削除
        like_num = re.sub(r"\D", "", like)

        # いいね数をリスト型に追加
        csvlist.append(like_num)

      else:
        # いいね数が表示されていない場合はリスト型に0を追加
        csvlist.append("0")

      # 求人のURLをリスト型に追加
      csvlist.append(elem_a.get_attribute('href'))
      
      # 設定したキーワードが求人タイトルに含まれるか判定
      if kw01 in elem_ttl.text:
        #キーワードが含まれる場合はリスト型にTrueを追加
        csvlist.append("True")
      
      else:
        #キーワードが含まれない場合はリスト型にFalseを追加
        csvlist.append("False")

      if kw02 in elem_ttl.text:
        csvlist.append("True")
      else:
        csvlist.append("False")

      if kw03 in elem_ttl.text:
        csvlist.append("True")
      else:
        csvlist.append("False")

      if kw04 in elem_ttl.text:
        csvlist.append("True")
      else:
        csvlist.append("False")

      if kw05 in elem_ttl.text:
        csvlist.append("True")
      else:
        csvlist.append("False")

      # csvファイルにリスト型の内容を追加
      writer.writerow(csvlist)

      item = item + 1

    # 次ページの有無を判定
    if driver.find_elements(By.LINK_TEXT, "次のページへ"):
      # 次ページが存在する場合は次ぺージへ遷移
      next_link = driver.find_element(By.LINK_TEXT, "次のページへ")
      driver.get(next_link.get_attribute('href'))

    else:
      # 次ページが存在しない場合はループ処理を終了
      break

    if i > 6:
      break

print ('csvファイルを作成しました。')

# 作成したcsvファイルをpandasで読み込み
df = pd.read_csv(csv_file_name, encoding="Shift-JIS")

# いいね数が多い順にソート
like_sort = df.sort_values("いいね数", ascending=False)

# デスクトップへのパスを取得
desktop_path = os.getenv("HOMEDRIVE") + os.getenv("HOMEPATH") + "\\Desktop"

# デスクトップにresultというフォルダを作成
dir_result = desktop_path + '\\' + 'result'
os.makedirs(dir_result, exist_ok=True)

# resultフォルダの中にファイル保存用のフォルダを作成
dir = dir_result + '\\' + csv_date
os.makedirs(dir, exist_ok=True)

# xlsxファイルをデスクトップのresultフォルダに書き出し
xlsx_file_name = 'engage' + csv_date + '.xlsx'
like_sort.to_excel(dir + '//' + xlsx_file_name, index=False)
print ('xlsxファイルを作成しました。')

# 作成したxlsxファイルの読み込み
inputfile = dir + '//' + xlsx_file_name
wb1 = xl.load_workbook(filename=inputfile)
ws1 = wb1.worksheets[0]

# 読み込んだシートの最終行を取得
max1 = ws1.max_row
print(max1)

# 黄緑色をセルに設定する処理
fill = PatternFill(patternType='solid', fgColor='0000FF00')

# Trueのセルを黄緑色に着色
for i in range(2, max1+1):
  if ws1['E' + str(i)].value == True:
    ws1['E' + str(i)].fill = fill
  
  if ws1['F' + str(i)].value == True:
    ws1['F' + str(i)].fill = fill

  if ws1['G' + str(i)].value == True:
    ws1['G' + str(i)].fill = fill

  if ws1['H' + str(i)].value == True:
    ws1['H' + str(i)].fill = fill

  if ws1['I' + str(i)].value == True:
    ws1['I' + str(i)].fill = fill


# ヒストグラムを作成するためのリスト型を定義
kw1_data = []
kw2_data = []
kw3_data = []
kw4_data = []
kw5_data = []


# キーワードが含まれている求人はそのいいね数順位をリスト型に格納
for i in range(2, max1+1):
  if ws1['E' + str(i)].value == True:
    kw1_data.append(str(i))

for i in range(2, max1+1):
  if ws1['F' + str(i)].value == True:
    kw2_data.append(str(i))

for i in range(2, max1+1):
  if ws1['G' + str(i)].value == True:
    kw3_data.append(str(i))

for i in range(2, max1+1):
  if ws1['H' + str(i)].value == True:
    kw4_data.append(str(i))

for i in range(2, max1+1):
  if ws1['I' + str(i)].value == True:
    kw5_data.append(str(i))

# figureとaxesを用意
fig = plt.figure()
ax1 = fig.add_subplot(3, 2, 1)
ax2 = fig.add_subplot(3, 2, 2)
ax3 = fig.add_subplot(3, 2, 3)
ax4 = fig.add_subplot(3, 2, 4)
ax5 = fig.add_subplot(3, 2, 5)

# ヒストグラムを作成
ax1.hist(kw1_data, bins=10, alpha=0.3)
ax2.hist(kw2_data, bins=10, alpha=0.3)
ax3.hist(kw3_data, bins=10, alpha=0.3)
ax4.hist(kw4_data, bins=10, alpha=0.3)
ax5.hist(kw5_data, bins=10, alpha=0.3)

ax1.set_title(kw01)
ax2.set_title(kw02)
ax3.set_title(kw03)
ax4.set_title(kw04)
ax5.set_title(kw05)

# ヒストグラムを画像として保存
plt.savefig(dir + '//' + "hst.png")

# xlsxファイルを保存
wb1.save(dir + '//' + xlsx_file_name)

# csvファイルはもう使わないので削除
# 作業ディレクトリにあるcsvファイルを全て指定
csv_files = glob.glob('*.csv')
for csv_file in csv_files:
  os.remove(csv_file)
print ('csvファイルを削除しました。')

# ブラウザを閉じる
driver.close()



